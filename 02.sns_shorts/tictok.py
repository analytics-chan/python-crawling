# 틱톡 쇼츠
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from webdriver_manager.chrome import ChromeDriverManager
from bs4 import BeautifulSoup

import time
from datetime import datetime, timedelta
import os
from openpyxl import Workbook, load_workbook

today = datetime.today()
year = str(today.year)
month = str(today.month)
day = str(today.day)

if len(month) == 1:
    month = '0' + month

if len(day) == 1:
    day = '0' + day

path = '02.sns_shorts'
file_list = os.listdir(path)
file_name = f'{year}{month}{day} tictok_shorts.xlsx'

if file_name in file_list:
    wb = load_workbook(os.path.join(path, file_name))
else:
    wb = Workbook()

    ws = wb.create_sheet('tictok', 0)

ws = wb.active

ws['A1'] = 'Date'
ws['B1'] = '계정ID'
ws['C1'] = '내용'
ws['D1'] = '작성일자'
ws['E1'] = '좋아요'
ws['F1'] = '댓글수'
ws['G1'] = '저장수'
ws['H1'] = '링크'

ws.column_dimensions['A'].width = 20
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 50
ws.column_dimensions['D'].width = 12
ws.column_dimensions['F'].width = 12

url = "https://www.tiktok.com/foryou"

chrome_options = Options()
chrome_options.add_experimental_option('detach', True)
chrome_options.add_experimental_option('excludeSwitches', ['enable-logging'])

service = Service(executable_path=ChromeDriverManager().install())
driver = webdriver.Chrome(service=service, options=chrome_options)

driver.implicitly_wait(5)
driver.maximize_window()
driver.get(url)

gue = driver.find_element(By.CSS_SELECTOR, '#loginContainer > div > div > div.css-txolmk-DivGuestModeContainer.exd0a435 > div > div > div > div > div')

time.sleep(1)

gue.click()


time.sleep(2)

vi = driver.find_element(By.CSS_SELECTOR, 'div.e1yey0rl2 > div > video')

vi.click()

time.sleep(1)

def info():
    id = driver.find_element(By.CSS_SELECTOR, 'span.css-1c7urt-SpanUniqueId.evv7pft1 > span').text
    # print(id)

    content = driver.find_element(By.CSS_SELECTOR, 'div.e1mecfx01').text

    wr_date = driver.find_element(By.CSS_SELECTOR, 'span.evv7pft3 > span:nth-child(3)').text

    like = driver.find_element(By.CSS_SELECTOR, 'div.ehlq8k31 > button:nth-child(1) > strong.e1hk3hf92').text

    save = driver.find_element(By.CSS_SELECTOR, 'div.ehlq8k31 > button:nth-child(3) > strong.e1hk3hf92').text

    review_count = driver.find_element(By.CSS_SELECTOR, 'div.e1aa9wve2').text

    # link = driver.find_element(By.CSS_SELECTOR, 'div.ehlq8k33 > p.ehlq8k34').text
    link = driver.current_url

    # now = datetime.now()

    # # 날짜 포맷 변경 함수
    # def convert_date_format(date_str):
    #     parts = date_str.split('-')
    #     month, day = map(int, parts)
    #     return now.replace(month=month, day=day).strftime("%Y-%m-%d")

    # ymd = convert_date_format(wr_date)
    # print(ymd)

    # # 날짜 데이터 전처리
    # if len(wr_date) >= 5:
    #     yymmdd = f"wr_date = \'{convert_date_format(wr_date)}\'"

    # if "시간전" in wr_date:
    #     hours = int(wr_date.split('시간전')[0])
    #     result_c = now - timedelta(hours=hours)
    #     yymmdd = f"wr_date = {result_c.strftime('%Y-%m-%d')}"

    # if "주전" in wr_date:
    #     weeks = int(wr_date.split('주전')[0])
    #     result_d = now - timedelta(weeks=weeks)
    #     yymmdd = f"wr_date = {result_d.strftime('%Y-%m-%d')}"


    print(id, content, wr_date, like, review_count[4:-1], save, link)
    ws.append([today.now(), id, content, wr_date, like, int(review_count[4:-1]), save, link])

    wb.save(os.path.join(path, file_name))

#app > div.css-14dcx2q-DivBodyContainer.e1irlpdw0 > div:nth-child(4) > div > div.css-1qjw4dg-DivContentContainer.e1mecfx00 > div.css-13if7zh-DivCommentContainer.ekjxngi0 > div > div.css-1xlna7p-DivProfileWrapper.ekjxngi4 > div.css-pcqxr7-DivDescriptionContentWrapper.e1mecfx011 > div.css-85dfh6-DivInfoContainer.evv7pft0 > a.evv7pft4.css-n2qh4e-StyledLink-StyledLink.er1vbsz0 > span.css-1c7urt-SpanUniqueId.evv7pft1 > span
#app > div.css-14dcx2q-DivBodyContainer.e1irlpdw0 > div:nth-child(4) > div > div.css-1qjw4dg-DivContentContainer.e1mecfx00 > div.css-13if7zh-DivCommentContainer.ekjxngi0 > div > div.css-1xlna7p-DivProfileWrapper.ekjxngi4 > div.css-pcqxr7-DivDescriptionContentWrapper.e1mecfx011 > div.css-85dfh6-DivInfoContainer.evv7pft0 > a.evv7pft4.css-n2qh4e-StyledLink-StyledLink.er1vbsz0 > span.css-1c7urt-SpanUniqueId.evv7pft1 > span

time.sleep(5)

info()

i = 0

while True:
    next = driver.find_element(By.CSS_SELECTOR, 'div.e11s2kul24 > button:nth-child(8)')

    next.click()

    time.sleep(2)

    info()

    i += 1

    time.sleep(2)

    print('i는 몇번째? >>> ', i)    

    if i >= 10:
        break

wb.save(os.path.join(path, file_name))