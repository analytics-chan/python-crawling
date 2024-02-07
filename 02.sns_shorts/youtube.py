from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from webdriver_manager.chrome import ChromeDriverManager
from bs4 import BeautifulSoup

import time
import datetime
import os
from openpyxl import Workbook, load_workbook

today = datetime.datetime.today()
year = str(today.year)
month = str(today.month)
day = str(today.day)

if len(month) == 1:
    month = '0' + month

if len(day) == 1:
    day = '0' + day

path = '02.sns_shorts'
file_list = os.listdir(path)
file_name = f'{year}{month}{day} youtube_shorts.xlsx'

if file_name in file_list:
    wb = load_workbook(os.path.join(path, file_name))
else:
    wb = Workbook()

    ws = wb.create_sheet('youtube', 0)

ws = wb.active

ws['A1'] = 'Date'
ws['B1'] = '제목'
ws['C1'] = '내용'
ws['D1'] = '작성일자'
ws['E1'] = '좋아요'
ws['F1'] = '댓글수'
ws['G1'] = '링크'

ws.column_dimensions['A'].width = 20
ws.column_dimensions['B'].width = 30
ws.column_dimensions['C'].width = 50
ws.column_dimensions['D'].width = 12

url = 'https://www.youtube.com/feed/trending?bp=6gQJRkVleHBsb3Jl'

chrome_options = Options()
chrome_options.add_experimental_option('detach', True)
chrome_options.add_experimental_option('excludeSwitches', ['enable-logging'])

service = Service(executable_path=ChromeDriverManager().install())
driver = webdriver.Chrome(service=service, options=chrome_options)

driver.implicitly_wait(5)
driver.maximize_window()
driver.get(url)

time.sleep(1)

top_shorts = driver.find_element(By.CSS_SELECTOR, 'div#scroll-container > div > ytd-reel-item-renderer > div > ytd-thumbnail > a')

top_shorts.click()

time.sleep(2)

lo_url = driver.current_url

etc = driver.find_element(By.CSS_SELECTOR, '#button-shape > button > yt-touch-feedback-shape > div')
etc.click()
time.sleep(1)

exp = driver.find_element(By.CSS_SELECTOR, '#items > ytd-menu-service-item-renderer > tp-yt-paper-item')
exp.click()
time.sleep(1)

etc2 = driver.find_element(By.CSS_SELECTOR, '#more > yt-formatted-string')
etc2.click()
time.sleep(1)


# div#container > div#text-container > yt-formatted-string#text > a.yt-simple-endpoint.style-scope.yt-formatted-string
# id = driver.find_element(By.CSS_SELECTOR, 'div#container > div#text-container > yt-formatted-string#text > a.yt-simple-endpoint.style-scope.yt-formatted-string').get_attribute('href')
# id = driver.find_element(By.CSS_SELECTOR, '//*[@id="text"]/a').text

def info():
    title = driver.find_element(By.CSS_SELECTOR, '#overlay > reel-player-header-renderer > h2 > yt-formatted-string').text
    content = driver.find_element(By.CSS_SELECTOR, '#description > yt-formatted-string').text
    like = driver.find_element(By.CSS_SELECTOR, '#like-button > yt-button-shape > label > div > span').text
    review_count = driver.find_element(By.CSS_SELECTOR, '#comments-button > ytd-button-renderer > yt-button-shape > label > div > span').text
    #description > yt-formatted-string

    # if "만" in review_count:
    #     rc_result = float(review_count.replace('만', '')) * 10000

    date_year = driver.find_element(By.CSS_SELECTOR, '#factoids > ytd-factoid-renderer:nth-child(2) > div > yt-formatted-string.factoid-label.style-scope.ytd-factoid-renderer').text
    date_md = driver.find_element(By.CSS_SELECTOR, '#factoids > ytd-factoid-renderer:nth-child(2) > div > yt-formatted-string.factoid-value.style-scope.ytd-factoid-renderer').text

    # 날짜 데이터 전처리
    ymd = date_year + ' ' + date_md
    ymd_r = ymd.replace('년', '-').replace('월', '-').replace('일', '')
    yymmdd = '-'.join(part.strip().zfill(2) for part in ymd_r.split('-'))

    print(yymmdd)

    print(title, content, yymmdd, like, review_count, lo_url)
    ws.append([today.now(), title, content, yymmdd, like, review_count, lo_url])
    # print(id)

    wb.save(os.path.join(path, file_name))


info()

i = 0

while True:
    try:
        next = driver.find_element(By.CSS_SELECTOR, '#navigation-button-down > ytd-button-renderer > yt-button-shape > button > yt-touch-feedback-shape > div > div.yt-spec-touch-feedback-shape__fill')
        # next.send_keys(Keys.PAGE_DOWN)
        next.click()

        time.sleep(2)

        info()

        i += 1

        time.sleep(5)
    except Exception as e:
        print('----END----')
        break

    if i >=100:
        break

# except Exception as e:
#     print('----END----')
#     break


wb.save(os.path.join(path, file_name))