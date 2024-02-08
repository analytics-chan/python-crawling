from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from webdriver_manager.chrome import ChromeDriverManager
from bs4 import BeautifulSoup

import time
import datetime
import os
from openpyxl import Workbook, load_workbook

today = datetime.datetime.today()
year = str(today.year)
month = str(today.month)
day = str(today.day)

if len(month) == 1:
    month = '0' + month

if len(day) == 1:
    day = '0' + day

path = '02.sns_shorts'
file_list = os.listdir(path)
file_name = f'{year}{month}{day} shorts_rank.xlsx'

if file_name in file_list:
    wb = load_workbook(os.path.join(path, file_name))
else:
    wb = Workbook()

    ws = wb.create_sheet('youtube', 0)

ws = wb.active

ws['A1'] = 'Date'
ws['B1'] = '채널명'
ws['C1'] = '구독자수'
ws['D1'] = '카테고리'
ws['E1'] = '작성일자'
ws['F1'] = '내용 및 해시태그'
ws['G1'] = '조회수'
ws['H1'] = '알고리즘 스코어'
ws['I1'] = '링크'

ws.column_dimensions['A'].width = 20
ws.column_dimensions['B'].width = 20
ws.column_dimensions['D'].width = 12
ws.column_dimensions['E'].width = 12
ws.column_dimensions['F'].width = 60
ws.column_dimensions['G'].width = 12
ws.column_dimensions['H'].width = 15

url = "https://vling.net/algorithm-score?na=KR&publishedAt=3&word="

chrome_options = Options()
chrome_options.add_experimental_option('detach', True)
chrome_options.add_experimental_option('excludeSwitches', ['enable-logging'])

service = Service(executable_path=ChromeDriverManager().install())
driver = webdriver.Chrome(service=service, options=chrome_options)

driver.implicitly_wait(5)
driver.maximize_window()
driver.get(url)

cards = driver.find_elements(By.CSS_SELECTOR, 'div.AlgoSearchTable_web__ovlWc')
print(len(cards))

for c in cards:
    channel_name = c.find_element(By.CSS_SELECTOR, 'p.ChannelInfoForColumn_title__OHUCG').text
    channel_score = c.find_element(By.CSS_SELECTOR, 'p.UnitNumber_number__8mgsP').text
    cate = c.find_element(By.CSS_SELECTOR, 'li.Category_category__eP9C9').text
    write_date = c.find_element(By.CSS_SELECTOR, 'p.TextForDate_date__sfjUl').text
    content = c.find_element(By.CSS_SELECTOR, 'p.VideoInfoInTable_title__K1eEu').text
    view_score = c.find_element(By.CSS_SELECTOR, 'p.NumericalText_numericalText__c7Y94').text
    al_score = c.find_element(By.CSS_SELECTOR, 'div.FormattedAlgoScoreText_wrapper__ZCtd3').text

    link = c.find_element(By.CSS_SELECTOR, 'div.VideoInfoInTable_wrapper__VabpU > a').get_attribute('href')
    link_result = 'https://www.youtube.com/shorts/' + link.split('/')[4]


    print(channel_name, channel_score, cate, write_date, content, view_score, al_score, link_result)
    ws.append([today.now(), channel_name, channel_score, cate, write_date, content, view_score, al_score, link_result])

    wb.save(os.path.join(path, file_name))

end_txt = f"{today.now()} 크롤링 완료"
ws.append([end_txt])
wb.save(os.path.join(path, file_name))