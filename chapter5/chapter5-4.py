# 광고 상품 제외하기
# 검색어 변경하기
# 상품 100개만 출력하기
# 엑셀에 저장하기
import requests
from bs4 import BeautifulSoup
import openpyxl

keyword = input('검색어를 입력하세요 >>> ')
page_num = int(input('몇 페이지까지 추출하시겠습니까? '))

wb = openpyxl.Workbook()
# wb = openpyxl.Workbook('chapter5/coupang_result.xlsx')
ws = wb.create_sheet(keyword, 0)
# ws = wb.create_sheet(keyword)
ws.append(['순위', '브랜드명', '상품명', '가격', '상세페이지링크'])

rank = 1
done = False

for i in range(1, page_num):
    # 이중 반복문 종료
    if done == True:
        break

    main_url = f"https://www.coupang.com/np/search?q={keyword}&channel=user&page={i}"

    # 헤더에 User-Agent를 추가하지 않으면 에러 발생(멈춤 현상)
    header = {
        'Host': 'www.coupang.com',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:76.0) Gecko/20100101 Firefox/76.0',
        'Accept' : 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
        'Accept-Language': 'ko-KR,ko;q=0.8,en-US;q=0.5,en;q=0.3',
    }

    response = requests.get(main_url, headers=header)
    html = response.text
    soup = BeautifulSoup(html, 'html.parser')

    # 결과는 리스트 자료형
    links = soup.select('a.search-product-link')
    # print(links)

    for link in links:
        # 광고 상품 제거
        if len(link.select('span.ad-badge-text')) > 0:
            print('광고 상품입니다.')
        else:
            sub_url = 'https://www.coupang.com/' + link.attrs['href']

            response = requests.get(sub_url, headers=header)
            html = response.text
            soup = BeautifulSoup(html, 'html.parser')

            # 브랜드명은 있을 수도 있고 없을 수도 있음
            # 중고상품일 경우 태그가 달라짐
            # try - except로 예외처리
            try:
                brand_name = soup.select_one('a.prod-brand-name ').text.strip()
            except:
                brand_name = ""

            # 상품명
            product_name = soup.select_one('h2.prod-buy-header__title').text

            # 가격
            try:
                price = soup.select_one('span.total-price > strong').text
            except:
                price = 0

            print(rank, brand_name, product_name, price)
            ws.append([rank, brand_name, product_name, price, sub_url])

            # time.sleep(0.3)

            rank += 1
            if rank > 100:
                done = True
                break

file_name = input('파일 이름을 입력하세요 >>> ')
# wb.save(rf'C:\Users\withbrother\Desktop\python_test2\chapter5\coupang_{file_name}.xlsx')
wb.save(f'chapter5/coupang_{file_name}.xlsx')
# wb.save('chapter5/coupang_result.xlsx')