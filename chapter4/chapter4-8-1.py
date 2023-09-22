# 네이버 뉴스 본문 링크 가져와서 엑셀에 저장하기
import requests
from bs4 import BeautifulSoup
import time

from openpyxl import Workbook
from openpyxl.styles import Alignment

# 검색어 키워드 가져오기
keyword = input('검색어를 입력하세요 >>> ')

# 여러 페이지 추출
lastPage = int(input('몇 페이지까지 추출하시겠습니까? '))

# 엑셀 생성하기
wb = Workbook()

# 엑셀 시트 생성하기
ws = wb.create_sheet(keyword, 0)

# 열 너비 조정
ws.column_dimensions['A'].width = 60
ws.column_dimensions['B'].width = 60
ws.column_dimensions['C'].width = 120

ws['A1'] = '기사 제목'
ws['B1'] = '기사 링크'
ws['C1'] = '기사 본문'

row = 2
for i in range(1, lastPage * 10, 10):
    response = requests.get(f"https://search.naver.com/search.naver?where=news&sm=tab_pge&query={keyword}&start={i}")
    html = response.text
    soup = BeautifulSoup(html, 'html.parser')

    articles = soup.select('div.info_group')

    for a in articles:
        links = a.select('a.info')

        if len(links) >= 2:
            url = links[1].attrs['href']

            response = requests.get(url)
            html = response.text
            soup = BeautifulSoup(html, 'html.parser')
            
            if 'entertain' in response.url:
                title = soup.select_one('.end_tit')
                content = soup.select_one('#articeBody')
            elif 'sports' in response.url:
                title = soup.select_one('h4.title')
                content = soup.select_one('#newsEndContents')

                divs = content.select('div')

                for div in divs:
                    div.decompose()
                
                paras = content.select('p')

                for para in paras:
                    para.decompose()
            else:
                title = soup.select_one('#title_area > span')
                content = soup.select_one('#dic_area')

            time.sleep(0.3)

            ws[f'A{row}'] = title.text.strip()
            ws[f'B{row}'] = url
            ws[f'C{row}'] = content.text.strip()[:1000]
            
            # 자동 줄바꿈
            ws[f'C{row}'].alignment = Alignment(wrap_text=True)

            row += 1

# # 셀 데이터 저장하기
# ws['A1'] = 'Coding'

# 엑셀 저장하기
fileName = input('파일 이름을 입력하세요 >>> ')
wb.save(rf'C:\Users\withbrother\Desktop\python_test2\chapter4\{fileName}.xlsx')