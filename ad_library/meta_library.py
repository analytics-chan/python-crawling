# 메타 광고 라이브러리
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from webdriver_manager.chrome import ChromeDriverManager
from bs4 import BeautifulSoup

import time
import datetime
import os
from openpyxl import Workbook, load_workbook

today = datetime.datetime.today()
year = str(today.year)
month = str(today.month)
day = str(today.day)

keyword = input('검색어를 입력해주세요 >>> ')

# url = "https://ko-kr.facebook.com/ads/library/?active_status=all&ad_type=all&country=KR&media_type=all"
url = f"https://ko-kr.facebook.com/ads/library/?active_status=all&ad_type=all&country=KR&q={keyword}&search_type=keyword_unordered&media_type=all"

path = 'ad_library'
file_list = os.listdir(path)
file_name = f'{year}{month}{day} 페이스북_광고라이브러리.xlsx'

if file_name in file_list:
    wb = load_workbook(os.path.join(path, file_name))
else:
    wb = Workbook()

    ws = wb.create_sheet('페이스북', -1)

ws = wb.active

ws.column_dimensions['A'].width = 20
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 9

ws.append(['Date', '광고주', '이미지 주소', '썸네일', '광고 내용'])

chrome_options = Options()
chrome_options.add_experimental_option('detach', True)
chrome_options.add_experimental_option('excludeSwitches', ['enable-logging'])

service = Service(executable_path=ChromeDriverManager().install())
driver = webdriver.Chrome(service=service, options=chrome_options)

driver.implicitly_wait(5)
driver.maximize_window()
driver.get(url)

last_height = driver.execute_script("return document.body.scrollHeight")
print('첫번째 before >>> ', last_height)

while True:
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight)")

    time.sleep(3)

    new_height = driver.execute_script("return document.body.scrollHeight")
    print('스크롤 후 >>> ', new_height)

    # if new_height == last_height:
    if new_height >= 5000:
        break
    last_height = new_height

    print('두번째 before >>> ', last_height)

lib_box = driver.find_elements(By.CSS_SELECTOR, 'div.x1dr75xp.xh8yej3.x16md763 > div > div.xh8yej3')
print('카드 수 >>> ', len(lib_box))

for i, lib in enumerate(lib_box, 1):
    # dat = lib.find_element(By.XPATH, f'//*[@id="content"]/div/div[1]/div/div[5]/div[2]/div[2]/div[4]/div[1]/div[{i}]/div/div[1]/div/div[1]/div[3]/span')
    dat = lib.find_element(By.CSS_SELECTOR, 'div.x3nfvp2.x1e56ztr > span.x8t9es0.xw23nyj.xo1l8bm.x63nzvj.x108nfp6.xq9mrsl.x1h4wwuj.xeuugli').text

    try:
        img = lib.find_element(By.CSS_SELECTOR, 'div.x1ywc1zp.x78zum5.xl56j7k.x1e56ztr.x1277o0a > img').get_attribute('src')
    except:
        img = '동영상 이미지'
    # print(dat.text)
        
    # try:
    #     title = lib.find_element(By.CSS_SELECTOR, 'div > div.xh8yej3 > div > div > div._7k71 > div > div > div > div > a > span').text
    # except:
    #     title = lib.find_element(By.CSS_SELECTOR, 'div > div.xh8yej3 > div > div > div._8nsi._8nqp._a25w > div > div > div > a > span').text

    title = lib.find_element(By.CSS_SELECTOR, 'span.x8t9es0.x1fvot60.xxio538.x108nfp6.xq9mrsl.x1h4wwuj.x117nqv4.xeuugli').text
    cont = lib.find_element(By.CSS_SELECTOR, 'div > div.xh8yej3 > div > div > div.x6ikm8r.x10wlt62 > div > span > div > div > div').text

    # print(dat, title, img, cont)
    # print(title, img, cont)
    print(f'{i}번째 글')
    ws.append([today.now(), title, img, '', cont])
    # print(dat)

wb.save(os.path.join(path, file_name))

# #content > div > div:nth-child(1) > div > div.x8bgqxi.x1n2onr6 > div._8n_0 > div:nth-child(2) > div.x1dr75xp.xh8yej3.x16md763 > div.xrvj5dj.xdq2opy.xexx8yu.xbxaen2.x18d9i69.xbbxn1n.xdoe023.xbumo9q.x143o31f.x7sq92a.x1crum5w > div:nth-child(1) > div > div.xh8yej3 > div > div > div.x6ikm8r.x10wlt62 > div > span > div > div > div > span
# #content > div > div:nth-child(1) > div > div.x8bgqxi.x1n2onr6 > div._8n_0 > div:nth-child(2) > div.x1dr75xp.xh8yej3.x16md763 > div.xrvj5dj.xdq2opy.xexx8yu.xbxaen2.x18d9i69.xbbxn1n.xdoe023.xbumo9q.x143o31f.x7sq92a.x1crum5w > div:nth-child(2) > div > div.xh8yej3 > div > div > div.x6ikm8r.x10wlt62 > div > span > div > div > div
    
