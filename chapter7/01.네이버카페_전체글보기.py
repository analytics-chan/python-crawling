# 네이버 카페 전체글보기 1000 페이지
# url input으로 받기
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE

import time
import datetime

today = datetime.datetime.today()
year = str(today.year)
month = str(today.month)
day = str(today.day)


url = input('------------------------------\n 카페 주소를 입력해주세요. \n (전체 URL을 입력해주세요 예: https://cafe.naver.com/카페ID) \n ------------------------------\n\n 입력 : ')

# "https://cafe.naver.com/navercafezz"

wb = openpyxl.Workbook()

ws1 = wb.create_sheet('source', 0)

ws1['B3'] = "크롤링 주소"
ws1['C3'] = url
ws1['B4'] = "옵션"
ws1['C4'] = "보기옵션 : 50개씩"
ws1['C5'] = "페이지 : 1000p"
ws1['C6'] = "전체글보기(가입인사/출석체크 제외)"

ws1.column_dimensions['B'].width = 15
ws1.column_dimensions['C'].width = 40

ws = wb.create_sheet('전체글보기', 0)

ws.column_dimensions['A'].width = 20
ws.column_dimensions['B'].width = 20
ws.column_dimensions['C'].width = 70
ws.column_dimensions['D'].width = 25
ws.column_dimensions['F'].width = 20
ws.column_dimensions['G'].width = 10

ws.append(['Date', '카테고리', '제목', '링크', '댓글수', '작성자', '작성일', '조회수'])

chrome_options = Options()
chrome_options.add_experimental_option('detach', True)
chrome_options.add_experimental_option('excludeSwitches', ['enable-logging'])

service = Service(executable_path=ChromeDriverManager().install())
driver = webdriver.Chrome(service=service, options=chrome_options)

driver.implicitly_wait(5)
driver.maximize_window()
driver.get(url)

# 전체글보기 클릭
driver.find_element(By.CSS_SELECTOR, '#menuLink0').click()

time.sleep(2)

driver.switch_to.frame('cafe_main')

# 50개씩 보기
driver.find_element(By.CSS_SELECTOR, '#listSizeSelectDiv > a').click()
driver.find_element(By.CSS_SELECTOR, '#listSizeSelectDiv > ul > li:nth-child(7) > a').click()

# 1 ~ 998페이지까지 크롤링 + 1000페이지까지 크롤링
for i in range(2, 1002):
# for i in range(2, 45):
    res = driver.page_source
    soup = BeautifulSoup(res, 'html.parser')

    article = soup.select('.article-board')
    trs = article[1].select('#main-area > div:nth-child(4) > table > tbody > tr')

    print(i - 1, '페이지-------------------')

    for tr in trs:
        try:
            cate = tr.select_one('a.link_name').text.strip()
            title = tr.select_one('a.article').text.strip()
            writer = tr.select_one('.p-nick > .m-tcol-c').text
            date = tr.select_one('.td_date').text
            view_count = tr.select_one('.td_view').text            
            link = tr.select_one('a.article').attrs['href']
        except:
            pass

        try:
            review_count = tr.select_one('a.cmt > em').text.strip()
        except:
            review_count = 0

        if date.find(':') != -1:
                date = f'{year}-{month}-{day}'
        elif date.find('.') != -1:
            date = date.replace('.', '-')[:-1]
        # print(date.find('.'))

        if "만" in view_count:
            vc_result = float(view_count.replace('만', '')) * 10000
        elif "," in view_count:
            vc_result = view_count.replace(',', '')
        else:
            vc_result = view_count

        t = title.replace('\n', '').replace('         ', '')
        t_result = ILLEGAL_CHARACTERS_RE.sub(r'', t)

        # 가입인사 카테고리 제외 + 출석체크도 제외
        if cate != '가입인사' and cate != '출석체크' and cate != '정회원 등업신청' and cate != '∞ 가입인사 1' and cate != '∞ 가입인사 2' and cate != '∞ 가입인사 3' and cate != '∞ 등업신청':
            # print(cate, title.replace('\n', '').replace('         ', ''), review_count, writer, date, view_count)
            # if date.find(':') != -1:
            #     date = f'{year}-{month}-{day}'
            # elif date.find('.') != -1:
            #     date = date.replace('.', '-')[:-1]
            # # print(date.find('.'))

            # ws.append([today.now(), cate, title.replace('\n', '').replace('         ', ''), url + link, int(review_count), writer, date, view_count])
            # ws.append([today.now(), cate, title.replace('\n', '').replace('         ', ''), url + link, int(review_count), writer, date, int(vc_result)])
            ws.append([today.now(), cate, t_result, url + link, int(review_count), writer, date, int(vc_result)])

            # 415 페이지에 조회쉬 1.2만 <- 숫자표기 에러
            # ws.append([cate, title.replace('\n', '').replace('         ', ''), int(review_count), writer, date, int(view_count)])

        # time.sleep(1)

    try:            
        if i%10 == 1:
            driver.find_element(By.CLASS_NAME, 'pgR').click()
        else:
            if i == 1000:
                i = '1,000'
                driver.find_element(By.LINK_TEXT, str(i)).click()
            else:
                driver.find_element(By.LINK_TEXT, str(i)).click()
    except Exception as e:
        print('----END----')
        break

    time.sleep(1)

wb.save(f'chapter7/{year}{month}{day} 네이버_전체글보기.xlsx')