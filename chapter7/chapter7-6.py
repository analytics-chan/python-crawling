# 네이버 쇼핑 검색, 연관검색어
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from webdriver_manager.chrome import ChromeDriverManager

import os
import time
import datetime
from openpyxl import Workbook, load_workbook
import requests

# ga4.sychoi@gmail.com 계정 테스트
# token = "xoxb-6162516044822-6154651620951-6Dq1Ua15rwfdKMaYgIzXJsGj"
token = 'xoxb-6162516044822-6154651620951-Zq7dWfuX6epWji7OqpNFczlV'
channel = "#create-error-test-bot"
text = "📢 에러 발생!!! 확인 바랍니다."

today = datetime.datetime.today()
year = str(today.year)
month = str(today.month)
day = str(today.day)

path = 'chapter7'
file_list = os.listdir(path)
file_name = f'{year}{month}{day} 네이버 쇼핑검색.xlsx'


if file_name in file_list:
    wb = load_workbook(os.path.join(path, file_name))

    ws = wb.active
else:
    wb = Workbook()

    ws = wb.create_sheet('네이버 쇼핑검색', 0)

ws['A1'] = 'Date'
ws['B1'] = '검색어'
ws['C1'] = '연관검색어'

ws.column_dimensions['A'].width = 18
ws.column_dimensions['B'].width = 12
ws.column_dimensions['C'].width = 30

url = "https://shopping.naver.com/home"

chrome_options = Options()
chrome_options.add_experimental_option('detach', True)
chrome_options.add_experimental_option('excludeSwitches', ['enable-logging'])

service = Service(executable_path=ChromeDriverManager().install())
driver = webdriver.Chrome(service=service, options=chrome_options)

driver.implicitly_wait(5)
driver.maximize_window()
driver.get(url)

try:
    # keyword = input('검색어를 입력하세요 >>> ')
    keyword = input('파일을 붙여넣기 해주세요 >>> ')

    wb2 = load_workbook(keyword)
    sheet = wb2.get_sheet_by_name('Sheet1')

    excel_key = []

    for i in range(1, 100):
        val = sheet[f'A{i}'].value

        if val != None:
            excel_key.append(val)
        else:
            break

    print(excel_key)

    for k in excel_key:
        input_query = driver.find_element(By.CSS_SELECTOR, 'input._searchInput_search_text_3CUDs')
        input_query.click()

        input_query.clear()

        time.sleep(1)

        input_query.send_keys(k)
        print(f'키워드 : {k}')
        # input_query.send_keys('비타민')

        # pyperclip.copy(keyword)
        # pyperclip.copy('비타민')
        # pyautogui.hotkey('ctrl', 'v')

        key_box = driver.find_elements(By.CSS_SELECTOR, 'div._autoComplete_basis_result_1cDj8._autoComplete_active_3_pom > div > ul')
        print(len(key_box))

        print('----- 검색창 연관검색어 -----')
        ws.append(['----- 검색창 연관검색어 -----'])

        time.sleep(1)
        
        if len(key_box) != 0:
            for i in range(1, len(key_box) + 1):
                lis = driver.find_elements(By.CSS_SELECTOR, f'#gnb-gnb > div._gnb_header_area_150KE > div > div._gnbLogo_gnb_logo_3eIAf > div > div._gnbSearch_gnb_search_3O1L2 > form > div._gnbSearch_inner_2Zksb > div:nth-child(2) > div > div._autoComplete_basis_result_1cDj8._autoComplete_active_3_pom > div > ul:nth-child({i}) > li')
                # print(len(lis))

                for l in lis:
                    try:
                        title = l.find_element(By.CSS_SELECTOR, 'em').text
                        # print(title)
                    except:
                        title = l.find_element(By.CSS_SELECTOR, 'li > a').text
                        # print(title)
                    
                    print(today.now(), k, title)
                    ws.append([today.now(), k, title])

                    # time.sleep(1)
        else:
            print('검색창 내 연관검색어가 없습니다.')

        time.sleep(1)

        search = driver.find_element(By.CSS_SELECTOR, 'button._searchInput_button_search_1n1aw')
        search.click()

        time.sleep(1)

        print('----- 쇼핑연관 검색어 -----')
        ws.append(['----- 쇼핑연관 검색어 -----'])

        shopping_keyword = driver.find_elements(By.CSS_SELECTOR, '#container > div.relatedTags_relation_tag__Ct0q2 > div > ul > li')

        if len(shopping_keyword) != 0:
            etc_btn = driver.find_element(By.CSS_SELECTOR, 'button.relatedTags_btn_more__Fdsm1')
            etc_btn.click()

            time.sleep(1)

            for s in shopping_keyword:
                shop_title = s.find_element(By.CSS_SELECTOR, 'li > a').text

                print(today.now(), k, shop_title)
                ws.append([today.now(), k, shop_title])

            # time.sleep(1)
        else:
            print('연관 검색어가 없습니다.')

        driver.back()

        time.sleep(1)

    print(f'----- {today.now()} 크롤링 종료 -----')
    ws.append([f'----- {str(today.now()).split(".")[0]} 크롤링 종료 -----'])

    wb.save(os.path.join(path, file_name))
except Exception as e:
# except:
    print(f'----- {text} -----')
    requests.post("https://slack.com/api/chat.postMessage",
    headers={"Authorization": "Bearer "+token},
    data={"channel": channel,"text": text})

    print(e)

### 코드 수정 필요함