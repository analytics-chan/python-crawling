# 구글 리뷰 크롤링
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from webdriver_manager.chrome import ChromeDriverManager
from bs4 import BeautifulSoup

import time
import datetime
import os
from openpyxl import Workbook, load_workbook

today = datetime.datetime.today()
year = str(today.year)
month = str(today.month)
day = str(today.day)

keyword = "더와이즈치과"

path = 'chapter7'
file_list = os.listdir(path)
file_name = f'{year}{month}{day} 구글_{keyword}_리뷰.xlsx'

if file_name in file_list:
    wb = load_workbook(os.path.join(path, file_name))
else:
    wb = Workbook()

    ws = wb.create_sheet('구글_리뷰', 0)

ws = wb.active

ws['A1'] = 'Date'
ws['B1'] = '신규여부'
ws['C1'] = '닉네임'
ws['D1'] = '지역가이드 여부'
ws['E1'] = '작성리뷰수'
ws['F1'] = '작성사진수'
ws['G1'] = '별점'
ws['H1'] = '작성일자'
ws['I1'] = '내용'

ws.column_dimensions['A'].width = 20
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 15


url = "https://www.google.com/search?q=%EB%8D%94%EC%99%80%EC%9D%B4%EC%A6%88%EC%B9%98%EA%B3%BC&oq=%EB%8D%94%EC%99%80&gs_lcrp=EgZjaHJvbWUqDAgAECMYJxiABBiKBTIMCAAQIxgnGIAEGIoFMg0IARAuGK8BGMcBGIAEMgYIAhBFGDkyDQgDEC4YrwEYxwEYgAQyBwgEEC4YgAQyBwgFEAAYgAQyBwgGEAAYgAQyBggHEEUYQagCALACAA&sourceid=chrome&ie=UTF-8"

chrome_options = Options()
chrome_options.add_experimental_option('detach', True)
chrome_options.add_experimental_option('excludeSwitches', ['enable-logging'])

service = Service(executable_path=ChromeDriverManager().install())
driver = webdriver.Chrome(service=service, options=chrome_options)

driver.implicitly_wait(5)
driver.maximize_window()
driver.get(url)

driver.find_element(By.CSS_SELECTOR, 'span.hqzQac > span > a').click()

# //*[@id="rhs"]/div[4]/div[3]/div/div[2]/div/div[2]/div/div/span[3]/span/a
# //*[@id="ow105"]/a

time.sleep(1)

lis = driver.find_elements(By.CSS_SELECTOR, 'div.WMbnJf.vY6njf.gws-localreviews__google-review')
before_len = len(lis)
print('before >>> ', before_len)

review_box = driver.find_element(By.CSS_SELECTOR, 'div.review-dialog-list')

try:
    while True:
        driver.execute_script('arguments[0].scrollBy(0, 2000)', review_box)

        time.sleep(1)

        lis = driver.find_elements(By.CSS_SELECTOR, 'div.WMbnJf.vY6njf.gws-localreviews__google-review')
        after_len = len(lis)
        print('after >>> ', after_len)

        time.sleep(1)

        if after_len == before_len:
            break

        before_len = after_len
except Exception as e:
    print('---- END -----')

try:
    review_list = driver.find_elements(By.CSS_SELECTOR, 'div.WMbnJf.vY6njf.gws-localreviews__google-review')
    print('총 리뷰 수 >>> ', len(review_list))

    for review in review_list:
        try:
            more_content = review.find_element(By.CSS_SELECTOR, 'a.review-more-link')
            more_content.click()
            print('클릭완료')

            time.sleep(1)
        except:
            pass
except:
    time.sleep(1)
    print('리뷰가 존재하지 않음')

res = driver.page_source
soup = BeautifulSoup(res, 'html.parser')

reviews = soup.select('div.WMbnJf.vY6njf.gws-localreviews__google-review')
print(len(reviews))

for r in reviews:
    nick = r.select_one('div.TSUbDb > a').text

    try:
        t = r.select_one('span.A503be').text

        t_re = t.split('·')
        # print(len(t_re))

        if len(t_re) == 3:
            power = t_re[0]
            re_cnt = t_re[1].split('리뷰 ')[1][:-1]
            ph_cnt = t_re[2].split('사진 ')[1][:-1]
        elif len(t_re) == 2:
            power = "-"
            re_cnt = t_re[0]
            ph_cnt = t_re[1].split('리뷰 ')[1][:-1]
        elif len(t_re) == 1:
            power = "-"
            re_cnt = t_re[0].split('리뷰 ')[1][:-1]
            ph_cnt = 0
    except:
        power = "-"
        re_cnt = 0
        ph_cnt = 0

    star = r.select_one('span.lTi8oc.z3HNkc').attrs['aria-label'].split('중 ')[1][:-3]
    wr_date = r.select_one('span.dehysf.lTi8oc').text

    try:
        new_re = r.select_one('span.dmZI8b.lTi8oc').text
    except:
        new_re = "-"
    
    try:
        try:
            comm = r.select_one('span.review-full-text').text
        except:
            comm = r.select_one('div.Jtu6Td > span > span').text
    except:
        comm = "-"

    # if ',' in ph_cnt:
    #     ph_cnt = ph_cnt.split(',').join()

    print(new_re, nick, power, re_cnt, ph_cnt, star, wr_date, comm)
    ws.append([today.now(), new_re, nick, power, re_cnt, ph_cnt, star, wr_date, comm])

end_txt = f"{today.now()} 크롤링 완료"
ws.append([end_txt])

wb.save(os.path.join(path, file_name))