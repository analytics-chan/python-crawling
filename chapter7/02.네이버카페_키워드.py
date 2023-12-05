# 네이버 카페 키워드 크롤링
# url input으로 받기
# 댓글내용으로 검색
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
from bs4 import BeautifulSoup

import time
import datetime
import os
from openpyxl import Workbook, load_workbook

today = datetime.datetime.today()
year = str(today.year)
month = str(today.month)
day = str(today.day)

url = input('------------------------------\n 카페 주소를 입력해주세요. \n (전체 URL을 입력해주세요 예: https://cafe.naver.com/카페ID) \n ------------------------------\n\n 입력 : ')
# url = "https://cafe.naver.com/navercafezz"

path = 'chapter7'
file_list = os.listdir(path)
file_name = f'{year}{month}{day} 네이버카페_키워드.xlsx'

if file_name in file_list:
    wb = load_workbook(os.path.join(path, file_name))
else:
    wb = Workbook()

    ws1 = wb.create_sheet('source', -1)

    ws1['B3'] = "크롤링 주소"
    ws1['C3'] = url
    ws1['B4'] = "옵션"
    ws1['C4'] = "보기옵션 : 50개씩"
    ws1['C5'] = "키워드별 검색 데이터"

    ws1.column_dimensions['B'].width = 15
    ws1.column_dimensions['C'].width = 40

    ws = wb.create_sheet('키워드', 0)

# wb = openpyxl.Workbook()
# wb = openpyxl.open(f'chapter7/{year}{month}{day} 키워드.xlsx')

ws = wb.active

# ws = wb.create_sheet('누네안과_키워드', 0)
# ws = wb.create_sheet('비엔빛_키워드', 0)

ws.column_dimensions['A'].width = 20
ws.column_dimensions['B'].width = 70
ws.column_dimensions['C'].width = 25
ws.column_dimensions['E'].width = 10
ws.column_dimensions['F'].width = 10

ws.append(['Date', '제목', '링크', '댓글수', '작성자', '작성일', '조회수', '키워드'])

chrome_options = Options()
chrome_options.add_experimental_option('detach', True)
chrome_options.add_experimental_option('excludeSwitches', ['enable-logging'])

service = Service(executable_path=ChromeDriverManager().install())
driver = webdriver.Chrome(service=service, options=chrome_options)

driver.implicitly_wait(5)
driver.maximize_window()
driver.get(url)

# noon = ['누네안과', 'ㄴㄴ안과', 'ㄴㄴ', '강남누네', '강남ㄴㄴ', '남양주누네', '남양주ㄴㄴ', '대구누네', '대구ㄴㄴ', '김안과', 'ㄱ안과', '한길안과', 'ㅎㄱ안과', '새빛안과', 'ㅅㅂ안과', '성모안과', 'ㅅㅁ안과']
# print(len(noon))
# bnb = ['비앤빛안과', 'ㅂㅇㅂ안과', 'ㅂㅇㅂ', '강남비앤빛', '강남ㅂㅇㅂ', '비앤비안과', '압구정안과', 'ㅇㄱㅈ안과', '서울밝은세상안과', 'ㅅㅇㅂㅇㅅㅅ안과', '밝은세상안과', 'ㅂㅇㅅㅅ안과', '눈에미소안과', 'ㄴㅇㅁㅅ안과', '에스앤유안과', 'ㅇㅅㅇㅇ안과', '아이리움안과', 'ㅇㅇㄹㅇ안과', '밝은눈안과', 'ㅂㅇㄴ안과', '누네빛안과', 'ㄴㄴㅂ안과', '파티마안과', 'ㅍㅌㅁ안과', '온누리안과', 'ㅇㄴㄹ안과', '밝은누리안과', 'ㅂㅇㄴㄹ안과', '메트로안과', 'ㅁㅌㄹ안과', '밝은눈안과', 'ㅂㅇㄴ안과', '스마일프로', '비쥬맥스800', 'visumax800']
# print(len(bnb))

noon = ['ㄴㄴ', 'ㄱㅇㄱ', '김안과', 'ㄱ안과', 'ㅅㅂ', '새빛안과', 'ㅅㅁ', '성모안과']

# search = driver.find_element(By.CSS_SELECTOR, '#topLayerQueryInput')
# search.click()

# search.send_keys(noon[0])
# # search.send_keys(bnb[n])

# driver.find_element(By.CSS_SELECTOR, '#cafe-search > form > .btn').click()

# time.sleep(2)

for n in range(0, len(noon), 1):
# for n in range(0, len(bnb), 1):
    print(noon[n])
    # print(bnb[n])
    
    # 검색창 클릭
    search = driver.find_element(By.CSS_SELECTOR, '#topLayerQueryInput')
    search.click()

    search.send_keys(noon[n])
    # search.send_keys(bnb[n])

    driver.find_element(By.CSS_SELECTOR, '#cafe-search > form > .btn').click()

    time.sleep(2)

    # 게시글 내용 크롤링
    driver.switch_to.frame('cafe_main')

    # search_input = driver.find_element(By.CSS_SELECTOR, '#query')
    # search_input.click()

    # search_input.send_keys(noon[n])

    driver.find_element(By.CSS_SELECTOR, '#currentSearchBy').click()

    time.sleep(1)

    driver.find_element(By.CSS_SELECTOR, '#divSearchBy > ul.select_list > li:nth-child(3) > a').click()

    time.sleep(1)

    driver.find_element(By.CSS_SELECTOR, 'div.list-search > form >  div.input_search_area > button.btn-search-green').click()

    time.sleep(1)    

    # driver.find_element(By.CSS_SELECTOR, '#listSizeSelectDiv > a').click()
    # driver.find_element(By.CSS_SELECTOR, '#listSizeSelectDiv > ul > li:nth-child(7) > a').click()

    page_len = len(driver.find_elements(By.CSS_SELECTOR, '.prev-next > a'))

    # print(page_len)

    if page_len == 0:
        print('등록된 게시물이 없습니다.')
    else:
        for i in range(2, 50):
            res = driver.page_source
            soup = BeautifulSoup(res, 'html.parser')

            trs = soup.select('#main-area > div:nth-child(5) > table > tbody > tr')

            print(i - 1, '페이지-------------------')

            for tr in trs:
                try:
                    title = tr.select_one('a.article').text.strip().replace('\n', '').replace('         ', '')
                    writer = tr.select_one('.p-nick > .m-tcol-c').text
                    date = tr.select_one('.td_date').text
                    view_count = tr.select_one('.td_view').text.replace(',', '')
                    link = tr.select_one('a.article').attrs['href']
                except:
                    pass
            
                try:
                    review_count = tr.select_one('a.cmt > em').text.strip()
                except:
                    review_count = 0

                if date.find(':') != -1:
                    date = f'{year}-{month}-{day}'
                elif date.find('.') != -1:
                    date = date.replace('.', '-')[:-1]

                if '만' in view_count:
                    view_count = float(view_count.replace('만', '')) * 10000

                # print(title, review_count, writer, date, view_count)

                ws.append([today.now(), title, url + link, int(review_count), writer, date, int(view_count), noon[n]])
                # ws.append([title, url + link, int(review_count), writer, date, int(view_count), bnb[n]])

                # time.sleep(1)
            try:
                if i%10 == 1:
                    driver.find_element(By.CLASS_NAME, 'pgR').click()
                else:
                    driver.find_element(By.LINK_TEXT, str(i)).click()
            except Exception as e:
                print('----END----')
                break

    driver.switch_to.default_content()

    # time.sleep(1)

wb.save(os.path.join(path, file_name))
# wb.save(f'chapter7/{year}{month}{day} 키워드.xlsx')
# wb.save('chapter7/20231016 키워드.xlsx')