# 네이버 카페 키워드 크롤링
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
from bs4 import BeautifulSoup
import openpyxl

import time
import datetime

today = datetime.datetime.today()
year = str(today.year)
month = str(today.month)
day = str(today.day)

# url = "https://cafe.naver.com/jihosoccer123"
url = "https://cafe.naver.com/soho"
# 카페 url input() 받기

wb = openpyxl.Workbook()
# wb = openpyxl.open(f'chapter7/{year}{month}{day} 키워드.xlsx')

ws1 = wb.create_sheet('source', 0)

ws1['B3'] = "크롤링 주소"
ws1['C3'] = url
# 카페 이름 input() 받기
ws1['C4'] = "셀러오션"
ws1['B5'] = "옵션"
ws1['C5'] = "보기옵션 : 50개씩"
ws1['C6'] = "보기옵션 : 제목만 보기"
ws1['C7'] = "키워드별 검색 데이터"

ws1.column_dimensions['B'].width = 15
ws1.column_dimensions['C'].width = 40

# 시트 이름 input() 받기
ws = wb.create_sheet('셀러오션_키워드검색어', 0)

ws.column_dimensions['A'].width = 19
ws.column_dimensions['B'].width = 70
ws.column_dimensions['C'].width = 25
ws.column_dimensions['E'].width = 25
ws.column_dimensions['F'].width = 10

ws.append(['Date', '제목', '링크', '댓글수', '작성자', '작성일', '조회수', '키워드'])

chrome_options = Options()
chrome_options.add_experimental_option('detach', True)
chrome_options.add_experimental_option('excludeSwitches', ['enable-logging'])

service = Service(executable_path=ChromeDriverManager().install())
driver = webdriver.Chrome(service=service, options=chrome_options)

driver.implicitly_wait(5)
driver.maximize_window()
driver.get(url)

# 키워드 input() 받아서 배열로 지정
key = ['인터넷', 'IPTV', '티비', 'CCTV', '와이파이']

for n in range(0, len(key), 1):
    print(key[n])
    
    # 검색창 클릭
    search = driver.find_element(By.CSS_SELECTOR, '#topLayerQueryInput')
    search.click()

    search.send_keys(key[n])

    driver.find_element(By.CSS_SELECTOR, '#cafe-search > form > .btn').click()

    time.sleep(2)

    driver.switch_to.frame('cafe_main')

    driver.find_element(By.CSS_SELECTOR, '#listSizeSelectDiv > a').click()
    driver.find_element(By.CSS_SELECTOR, '#listSizeSelectDiv > ul > li:nth-child(7) > a').click()

    page_len = len(driver.find_elements(By.CSS_SELECTOR, '.prev-next > a'))

    if page_len == 0:
        print('등록된 게시물이 없습니다.')
    else:
        for i in range(2, 100):
            res = driver.page_source
            soup = BeautifulSoup(res, 'html.parser')

            trs = soup.select('#main-area > div:nth-child(5) > table > tbody > tr')

            print(i - 1, '페이지-------------------')

            for tr in trs:
                try:
                    title = tr.select_one('a.article').text.strip().replace('\n', '').replace('         ', '')
                    writer = tr.select_one('.p-nick > .m-tcol-c').text
                    date = tr.select_one('.td_date').text
                    view_count = tr.select_one('.td_view').text.replace(',', '')
                    link = tr.select_one('a.article').attrs['href']
                except:
                    pass
            
                try:
                    review_count = tr.select_one('a.cmt > em').text.strip()
                except:
                    review_count = 0

                if date.find(':') != -1:
                    date = f'{year}-{month}-{day}'
                elif date.find('.') != -1:
                    date = date.replace('.', '-')[:-1]

                if '만' in view_count:
                    view_count = float(view_count.replace('만', '')) * 10000

                # print(title, review_count, writer, date, view_count)

                ws.append([today.now(), title, url + link, int(review_count), writer, date, int(view_count), key[n]])
                # ws.append([title, url + link, int(review_count), writer, date, int(view_count), bnb[n]])

                # time.sleep(1)
            try:
                if i%10 == 1:
                    driver.find_element(By.CLASS_NAME, 'pgR').click()
                else:
                    driver.find_element(By.LINK_TEXT, str(i)).click()
            except Exception as e:
                print('----END----')
                break

    driver.switch_to.default_content()

    # time.sleep(1)

# 파일 이름 input() 받기
wb.save(f'chapter7/{year}{month}{day} 셀러오션_키워드.xlsx')
# wb.save('chapter7/20231016 키워드.xlsx')