from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from webdriver_manager.chrome import ChromeDriverManager
from bs4 import BeautifulSoup

import time
import datetime
import os
from openpyxl import Workbook, load_workbook

today = datetime.datetime.today()
year = str(today.year)
month = str(today.month)
day = str(today.day)

# keyword = input('검색어를 입력하세요 >>> ')
keyword = "더와이즈치과병원"

path = 'chapter7'
file_list = os.listdir(path)
file_name = f'{year}{month}{day} {keyword}_리뷰.xlsx'

if file_name in file_list:
    wb = load_workbook(os.path.join(path, file_name))
else:
    wb = Workbook()

    ws = wb.create_sheet('네이버플레이스_리뷰', 0)

ws = wb.active

ws['A1'] = 'Date'
ws['B1'] = '닉네임'
ws['C1'] = '작성리뷰수'
ws['D1'] = '작성일자'
ws['E1'] = '방문수'
ws['F1'] = '인증수단'
ws['G1'] = '내용'

ws.column_dimensions['A'].width = 20
ws.column_dimensions['B'].width = 20
ws.column_dimensions['C'].width = 10
ws.column_dimensions['D'].width = 11
ws.column_dimensions['F'].width = 10

url = f"https://map.naver.com/p/search/{keyword}"
# url = f"https://map.naver.com/p/search/{keyword}/place/33084820?c=15.00,0,0,0,dh&isCorrectAnswer=true"

chrome_options = Options()
chrome_options.add_experimental_option('detach', True)
chrome_options.add_experimental_option('excludeSwitches', ['enable-logging'])

service = Service(executable_path=ChromeDriverManager().install())
driver = webdriver.Chrome(service=service, options=chrome_options)

driver.implicitly_wait(5)
driver.maximize_window()
driver.get(url)

# driver.find_element(By.TAG_NAME, 'body').send_keys(Keys.PAGE_DOWN)
time.sleep(2)

driver.switch_to.frame("entryIframe")

#iframe 밖으로 나오기
#browser.switch_to.default_content()

# driver.find_element(By.CSS_SELECTOR, 'body').send_keys(Keys.END)

time.sleep(2)
# review = driver.find_element(By.CSS_SELECTOR, '#app-root > div > div > div > div.place_fixed_maintab > div > div > div > div > a:nth-child(2) > span')
review = driver.find_element(By.CSS_SELECTOR, '#app-root > div > div > div > div.place_fixed_maintab > div > div > div > div > a:nth-child(2)')
# review = driver.find_element(By.XPATH, '//*[@id="app-root"]/div/div/div/div[5]/div/div/div/div/a[2]')
review.click()

lis = driver.find_elements(By.CSS_SELECTOR, 'li.YeINN')
before_len = len(lis)
print(before_len)

#app-root > div > div > div > div.place_fixed_maintab > div > div > div > div > a:nth-child(2)

# i = 1

try:
    while True:
        driver.find_element(By.CSS_SELECTOR, 'body').send_keys(Keys.END)

        time.sleep(1)
        
        # driver.find_element(By.XPATH, '//*[@id="app-root"]/div/div/div/div[6]/div[3]/div[3]/div[2]/a').click()
        driver.find_element(By.CSS_SELECTOR, '#app-root > div > div > div > div:nth-child(6) > div:nth-child(3) > div.place_section.k5tcc > div.NSTUp > div > a > span').click()

        time.sleep(1)

        lis = driver.find_elements(By.CSS_SELECTOR, 'li.YeINN')
        after_len = len(lis)
        print(after_len)
        
        time.sleep(1)

        # if after_len == 30:
        #     break

        if after_len == before_len:
            break

        before_len = after_len
        print(before_len)

        # if i == 10:
        #     break
except Exception as e:
    print('---END---')

# res = driver.page_source
# soup = BeautifulSoup(res, 'html.parser')

try:
    review_list = driver.find_elements(By.CSS_SELECTOR, 'li.YeINN')
    
    print('총 리뷰 수 : ', len(review_list))

    for review in review_list:
        try:
            # find_review = review.find('div.ZZ4OK > a > span.rvCSr > svg')
            # print(find_review)
            more_content = review.find_element(By.CSS_SELECTOR, 'div.ZZ4OK > a > span.rvCSr > svg')
            # print(more_content)
            more_content.click()
            print('클릭완료')

            time.sleep(1)

            # user_review = review.find_element(By.CSS_SELECTOR, 'div.ZZ4OK > a > span.zPfVt')
        except:
            # user_review = review.find_element(By.CSS_SELECTOR, 'div.ZZ4OK > a > span.zPfVt')
            pass
            # time.sleep(1)
except:
    time.sleep(1)
    print('리뷰가 존재하지 않음')

driver.find_element(By.CSS_SELECTOR, 'body').send_keys(Keys.END)

res = driver.page_source
soup = BeautifulSoup(res, 'html.parser')

reviews = soup.select('li.YeINN')
print(len(reviews))

for review in reviews:
    nick = review.select_one('div.VYGLG').text

    try:
        review_cnt = review.select_one('span.Eu5rp').text[2:].strip()
    except:
        review_cnt = 0
    
    try:
        user_review = review.select_one('div.ZZ4OK > a > span.zPfVt').text.replace('\n', ' ')
    except:
        user_review = ""

    try:
        ymd = review.select_one('div._7kR3e > span:nth-child(1) > span:nth-child(3)').text[:-4]

        yy = ymd.split('년')[0].strip()
        mm = ymd.split('년')[1].split('월')[0].strip()
        dd = ymd.split('년')[1].split('월')[1][:-1].strip()
        
        if len(mm) == 1:
            mm = '0' + mm
        
        if len(dd) == 1:
            dd = '0' + dd

        write_date = f'{yy}-{mm}-{dd}'
    except:
        write_date = ""

    try:
        visit = review.select_one('div._7kR3e > span:nth-child(2)').text[:-5]
    except:
        visit = ""

    try:
        cert = review.select_one('div._7kR3e > span:nth-child(3)').text[5:]
    except:
        cert = ""

    print(nick, review_cnt, write_date, visit, cert, user_review)
    ws.append([today.now(), nick, int(review_cnt), write_date, int(visit), cert, user_review])

end_txt = f"{today.now()} 크롤링 완료"
ws.append([end_txt])

wb.save(os.path.join(path, file_name))

