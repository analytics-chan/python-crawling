import requests
from bs4 import BeautifulSoup
import openpyxl


codes = [
  '005930',
  '000660',
  '035720'
]

wb = openpyxl.Workbook()

ws = wb.create_sheet('주식현재가')

# 현재 활성화된 시트 선택
# ws = wb.active

ws['A1'] = '종목'
ws['B1'] = '현재가'
ws['C1'] = '평균매입가'
ws['D1'] = '잔고수량'
ws['E1'] = '평가금액'
ws['F1'] = '평가손익'
ws['G1'] = '수익률'

ws['C2'] = 85000
ws['C3'] = 120000
ws['C4'] = 145000
ws['D2'] = 20
ws['D3'] = 15
ws['D4'] = 10
# ws['E2', 'E3', 'E4'] = '-'
# ws['F2', 'F3', 'F4'] = '######'
# ws['G2', 'G3', 'G4'] = '-100%'

row = 2

for code in codes:
  response = requests.get(f'https://finance.naver.com/item/sise.naver?code={code}')
  html = response.text
  soup = BeautifulSoup(html, 'html.parser')
  title = soup.select_one('#middle > div.h_company > div.wrap_company > h2 > a').text
  price = soup.select_one('#_nowVal').text
  price = price.replace(',', '')
  # print(title, price)
  ws[f'B{row}'] = int(price)
  ws[f'A{row}'] = title
  row = row + 1

wb.save(r'C:\Users\withbrother\Desktop\python_test2\chapter2\주식현재가.xlsx')