import requests
from bs4 import BeautifulSoup
import openpyxl

fpath = r'C:\Users\withbrother\Desktop\python_test2\chapter2\TOP종목_상한가.xlsx'

# wb = openpyxl.Workbook()

# ws = wb.create_sheet('TOP종목_상한가')

# ws['A1'] = '종목명'
# ws['B1'] = '종목코드'
# ws['C1'] = '현재가'
# ws['D1'] = '거래량'
# ws['E1'] = '시가'
# ws['F1'] = '고가'
# ws['G1'] = '저가'

row = 2
response = requests.get('https://finance.naver.com/sise/')
html = response.text
soup = BeautifulSoup(html, 'html.parser')
title = soup.select_one('#siselist_tab_0 > tbody > tr:nth-child(3) > td:nth-child(4) > a')
# title = soup.select('.title')

#siselist_tab_0 > tbody > tr:nth-child(3) > td:nth-child(4) > a
#siselist_tab_0 > tbody > tr:nth-child(4) > td:nth-child(4) > a

# print(title)
print(soup)

# wb.save(fpath)