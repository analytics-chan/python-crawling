import openpyxl

fpath = r'C:\Users\withbrother\Desktop\python_test2\chapter2\파이썬_엑셀만들기_test.xlsx'

# 1. 엑셀 불러오기
wb = openpyxl.load_workbook(fpath)

# 2. 엑셀 시트 선택
ws = wb['파이썬_엑셀만들기']

# 3. 데이터 수정하기
ws['A3'] = 456
ws['B3'] = '파이썬'

# 4. 엑셀 저장하기
wb.save(fpath)