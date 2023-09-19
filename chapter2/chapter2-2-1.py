import openpyxl

# 1. 엑셀 만들기
wb = openpyxl.Workbook()

# 2. 엑셀 워크시트 만들기
ws = wb.create_sheet('파이썬_엑셀만들기')

# 3. 데이터 추가하기
ws['A1'] = '참가번호'
ws['B1'] = '성명'

ws['A2'] = 1
ws['B2'] = '테스트'

# 4. 엑셀 저장하기
wb.save(r'C:\Users\withbrother\Desktop\python_test2\chapter2\파이썬_엑셀만들기_test.xlsx')